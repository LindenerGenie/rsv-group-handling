from flask import Flask, send_file, send_from_directory, jsonify, request
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import pandas as pd
import logging
import os

app = Flask(__name__, static_folder='dist')
CORS(app)

users_data = []
column_order = []

@app.route('/upload', methods=['POST'])
def upload_xlsx():
    if 'file' not in request.files:
        return 'No file part', 400
    file = request.files['file']
    if file.filename == '':
        return 'No selected file', 400
    wb = load_workbook(file)
    ws = wb.active

    # Initialize column_order with the first row which contains the column names
    global column_order
    column_order = [cell.value for cell in ws[1]]

    global users_data
    users_data = []
    # Include 'created' and 'groups' in the expected keys
    default_keys = ['firstname', 'lastname', 'email', 'departments', 'created', 'groups']  # Updated expected keys

    # Read the first row to get column headers
    headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    # Create a mapping of headers to their index
    header_index = {header: index for index, header in enumerate(headers)}

    for row in ws.iter_rows(min_row=2, values_only=True):
        user = {key: '' for key in default_keys}  # Initialize with default keys and empty values

        # Update user dictionary with values from row, using header_index for correct mapping
        for key in default_keys:
            if key in header_index:  # Check if the key exists in the header
                value_index = header_index[key]
                user[key] = row[value_index] if value_index < len(row) else ''

        users_data.append(user)
    return jsonify(success=True), 200

@app.route('/users', methods=['GET'])
def get_users():
    search = request.args.get('search', '').lower()
    department = request.args.get('department', '')
    filtered_users = []
    for user in users_data:
        # Use .get() with default empty string to avoid KeyError
        lastname = user.get('lastname', '').lower()
        firstname = user.get('firstname', '').lower()
        email = user.get('email', '').lower()
        departments = user.get('departments', '').split(',')
        if (search in lastname or
            search in firstname or
            search in email or
            (department and department in departments)):
            filtered_users.append(user)
    return jsonify(filtered_users)

@app.route('/update-groups', methods=['POST'])
def update_groups():
    data = request.json
    user_ids = data['userIds']
    groups_to_add = data['groupsToAdd']
    groups_to_remove = data['groupsToRemove']

    for user in users_data:
        user_id = f"{user['firstname']}_{user['lastname']}_{user['email']}"
        if user_id in user_ids:
            current_groups = set(user.get('groups', '').split(';')) if user.get('groups') else set()
            current_groups.update(groups_to_add)
            current_groups.difference_update(groups_to_remove)
            user['groups'] = ';'.join(sorted(current_groups))

    return jsonify({"message": "Groups updated successfully"})

@app.route('/export', methods=['GET'])
def export_xlsx():
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(column_order)  # Assuming column_order contains the headers
        filtered_users = [user for user in users_data if user.get('firstname') and user.get('lastname') and user.get('email')]
        for user in filtered_users:
            # Ensure data is correctly formatted for Excel
            row = [user.get(col, '') for col in column_order]
            ws.append(row)
        filename = "exported_users.xlsx"
        wb.save(filename)
        return send_file(filename, as_attachment=True, attachment_filename=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        logging.error("An error occurred while exporting users: %s", str(e))  # Log the exception
        return jsonify({"error": "An error occurred while exporting users."}), 500

@app.route('/groups', methods=['GET'])
def get_groups():
    all_groups = set()
    for user in users_data:
        if user.get('groups'):
            all_groups.update(user['groups'].split(';'))
    return jsonify(list(all_groups))

@app.route('/users')
def users():
    # Adjust the path to your XLSX file
    file_path = 'path/to/your/users.xlsx'
    # Read the XLSX file into a DataFrame
    df = pd.read_excel(file_path)
    # Convert the DataFrame to a list of dictionaries for JSON response
    data = df.to_dict(orient='records')
    return jsonify(data)

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve(path):
    if path != "" and os.path.exists(app.static_folder + '/' + path):
        return send_from_directory(app.static_folder, path)
    else:
        return send_from_directory(app.static_folder, 'index.html')

@app.route('/test')
def test():
    return "Hello, World!"

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
